"""
mapper.py — Maps extracted data to Excel cells using BOLD+RED expressions.
"""

import re
import json
import time
import openpyxl
from typing import Any, List
from .config import ProcessingContext


class Mapper:
    def __init__(self, context: ProcessingContext):
        self.context = context

    def _normalize_name(self, s: str) -> str:
        return re.sub(r"[^0-9a-zA-Z]", "", s).lower()

    def resolve_expression(self, expr: str) -> Any:
        # Simple resolution logic similar to original, but using context dict
        data_map = self.context.to_dict()

        expr = expr.strip()
        m = re.match(r"^([A-Za-z_]\w*)(.*)$", expr)
        if not m:
            return None

        base_raw = m.group(1)
        rest = m.group(2) or ""

        # Normalize base key lookup
        base_key = None
        norm_map = {self._normalize_name(k): k for k in data_map.keys()}
        base_key = norm_map.get(self._normalize_name(base_raw))

        if not base_key:
            return None

        current_obj = data_map[base_key]

        # Parse brackets: ['key'] or ["key"]
        keys = re.findall(r"\[['\"](.*?)['\"]\]", rest)

        for k in keys:
            if isinstance(current_obj, dict):
                # Case insensitive lookup
                found = None
                for real_k in current_obj.keys():
                    if self._normalize_name(real_k) == self._normalize_name(k):
                        found = real_k
                        break
                if found:
                    current_obj = current_obj[found]
                else:
                    return None
            else:
                return None

        return current_obj

    def map_to_excel(self, xlsx_path: str):
        start = time.time()
        self.context.log("[MAPPER] ▶ Scanning workbook for BOLD+RED expressions...")
        
        try:
            wb = openpyxl.load_workbook(xlsx_path)
            sheet = wb.active
            self.context.log(f"[MAPPER]   Workbook loaded: {sheet.max_row} rows × {sheet.max_column} cols")

            cells_to_update = []
            bold_red_found = 0
            resolution_failures = []

            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=20):
                for cell in row:
                    val = cell.value
                    if not val or not isinstance(val, str):
                        continue

                    # Check for Red Bold
                    font = cell.font
                    if not font or not font.bold:
                        continue
                    col = font.color
                    if not col or not col.rgb or (isinstance(col.rgb, str) and str(col.rgb).upper()[-6:] != "FF0000"):
                        continue

                    bold_red_found += 1

                    # Clean expression
                    expr = val.strip()
                    if (expr.startswith('"') and expr.endswith('"')) or (expr.startswith("'") and expr.endswith("'")):
                        expr = expr[1:-1].strip()

                    # Resolve
                    resolved_val = self.resolve_expression(expr)

                    if resolved_val is not None:
                        cells_to_update.append((cell, resolved_val, expr))
                    else:
                        resolution_failures.append((cell.coordinate, expr))
            
            self.context.log(f"[MAPPER]   Found {bold_red_found} BOLD+RED cells")
            self.context.log(f"[MAPPER]   Resolved: {len(cells_to_update)} / Unresolved: {len(resolution_failures)}")

            # Log unresolved expressions for debugging
            if resolution_failures:
                self.context.log(f"[MAPPER]   ⚠ Unresolved expressions:")
                for coord, expr in resolution_failures[:10]:  # Cap at 10
                    self.context.log(f"[MAPPER]     Cell {coord}: \"{expr}\"")
                if len(resolution_failures) > 10:
                    self.context.log(f"[MAPPER]     ... and {len(resolution_failures) - 10} more")

            # Apply updates
            count = 0
            for cell, val, expr in cells_to_update:
                old_val = cell.value
                if isinstance(val, (dict, list)):
                    cell.value = json.dumps(val)
                else:
                    cell.value = val
                count += 1
                self.context.log(f"[MAPPER]   ✓ {cell.coordinate}: \"{expr}\" → {repr(cell.value)[:80]}")

            wb.save(xlsx_path)
            elapsed = time.time() - start
            self.context.log(f"[MAPPER] ✅ Updated {count} cells, saved workbook ({elapsed:.1f}s)")

        except Exception as e:
            elapsed = time.time() - start
            self.context.log(f"[MAPPER] ✗ MAPPING FAILED: {type(e).__name__}: {e} ({elapsed:.1f}s)")
            import traceback
            self.context.log(f"[MAPPER]   Traceback: {traceback.format_exc()}")
